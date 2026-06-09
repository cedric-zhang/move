"""
Excel 导出器测试
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.excel_exporter import ExcelExporter
from openpyxl import load_workbook
from io import BytesIO


def test_generate_has_two_sheets():
    """测试生成两个 Sheet"""
    creds = [
        {"id": 1, "type": "SNMPv2", "name": "public", "credential_value": "public", "field": "community", "host_count": 2, "host_list": "Host1, Host2"},
        {"id": 2, "type": "SNMPv3", "name": "auth", "credential_value": "authpass", "field": "auth_passphrase", "host_count": 1, "host_list": "Host3"},
    ]
    summary = [
        {"type": "SNMPv2", "count": 1, "host_count": 2},
        {"type": "SNMPv3", "count": 1, "host_count": 1},
    ]

    exporter = ExcelExporter(creds, summary)
    excel_bytes = exporter.generate()

    wb = load_workbook(BytesIO(excel_bytes))
    assert len(wb.sheetnames) == 2
    assert "凭证清单" in wb.sheetnames
    assert "按类型统计" in wb.sheetnames

    print(f"[OK] test_generate_has_two_sheets passed -- sheets: {wb.sheetnames}")


def test_sheet1_has_headers():
    """测试 Sheet1 有正确表头"""
    creds = [{"id": 1, "type": "SNMPv2", "name": "public", "credential_value": "public", "field": "community", "host_count": 1, "host_list": "Host1"}]
    summary = [{"type": "SNMPv2", "count": 1, "host_count": 1}]

    exporter = ExcelExporter(creds, summary)
    excel_bytes = exporter.generate()

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb["凭证清单"]

    headers = [ws.cell(row=1, column=i).value for i in range(1, 8)]
    expected = ["ID", "类型", "名称", "认证方式", "凭据值", "使用主机数", "主机列表"]
    assert headers == expected

    print(f"[OK] test_sheet1_has_headers passed -- headers: {headers}")


def test_sheet2_has_summary():
    """测试 Sheet2 有统计数据"""
    creds = [{"id": 1, "type": "SNMPv2", "name": "public", "credential_value": "public", "field": "community", "host_count": 1, "host_list": "Host1"}]
    summary = [{"type": "SNMPv2", "count": 1, "host_count": 1}, {"type": "SSH", "count": 2, "host_count": 3}]

    exporter = ExcelExporter(creds, summary)
    excel_bytes = exporter.generate()

    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb["按类型统计"]

    # 第一行是表头
    assert ws.cell(row=1, column=1).value == "类型"
    assert ws.cell(row=1, column=2).value == "数量"

    # 第二行是数据
    assert ws.cell(row=2, column=1).value == "SNMPv2"
    assert ws.cell(row=2, column=2).value == 1

    print("[OK] test_sheet2_has_summary passed")


def test_export_bytes_not_empty():
    """测试导出字节不为空"""
    creds = [{"id": 1, "type": "SNMPv2", "name": "public", "credential_value": "public", "field": "community", "host_count": 1, "host_list": "Host1"}]
    summary = [{"type": "SNMPv2", "count": 1, "host_count": 1}]

    exporter = ExcelExporter(creds, summary)
    excel_bytes = exporter.generate()

    assert len(excel_bytes) > 0
    assert excel_bytes[:4] == b"PK\x03\x04"  # xlsx 文件开头

    print(f"[OK] test_export_bytes_not_empty passed -- size: {len(excel_bytes)} bytes")


def run_all_tests():
    """运行所有测试"""
    print("=== Excel Exporter Tests ===")
    test_generate_has_two_sheets()
    test_sheet1_has_headers()
    test_sheet2_has_summary()
    test_export_bytes_not_empty()
    print("=== All tests passed! ===")


if __name__ == "__main__":
    run_all_tests()