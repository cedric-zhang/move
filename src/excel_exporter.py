"""
Excel 导出器 - 生成凭证 Excel 文件
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime


class ExcelExporter:
    """生成凭证 Excel 文件"""

    # 样式定义
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="5E6AD2", end_color="5E6AD2", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    BORDER = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    def __init__(self, credentials: List[Dict[str, Any]], summary: List[Dict[str, Any]] = None):
        """传入凭证列表和统计信息"""
        self.credentials = credentials
        self.summary = summary or []

    def generate(self) -> bytes:
        """
        生成双 Sheet xlsx：
        Sheet1 "凭证清单": ID | 类型 | 名称 | 认证方式 | 凭据值 | 使用主机数 | 主机列表
        Sheet2 "按类型统计": 类型 | 数量 | 涉及主机数
        返回 xlsx 文件字节
        """
        wb = Workbook()

        # Sheet1: 凭证清单
        ws1 = wb.active
        ws1.title = "凭证清单"
        self._write_sheet1(ws1)

        # Sheet2: 按类型统计
        ws2 = wb.create_sheet(title="按类型统计")
        self._write_sheet2(ws2)

        # 导出为字节
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_sheet1(self, ws):
        """写入凭证清单 Sheet"""
        headers = ["ID", "类型", "名称", "认证方式", "凭据值", "使用主机数", "主机列表"]
        col_widths = [8, 12, 16, 18, 20, 12, 40]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

        # 写入数据
        for row, cred in enumerate(self.credentials, 2):
            ws.cell(row=row, column=1, value=cred.get("id", row - 1)).border = self.BORDER
            ws.cell(row=row, column=2, value=cred.get("type", "")).border = self.BORDER
            ws.cell(row=row, column=3, value=cred.get("name", "")).border = self.BORDER
            ws.cell(row=row, column=4, value=cred.get("field", "")).border = self.BORDER
            ws.cell(row=row, column=5, value=cred.get("credential_value", "")).border = self.BORDER
            ws.cell(row=row, column=6, value=cred.get("host_count", 0)).border = self.BORDER
            ws.cell(row=row, column=7, value=cred.get("host_list", "")).border = self.BORDER

    def _write_sheet2(self, ws):
        """写入统计 Sheet"""
        headers = ["类型", "数量", "涉及主机数"]
        col_widths = [15, 10, 12]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER
            ws.column_dimensions[get_column_letter(col)].width = col_widths[col - 1]

        # 写入统计数据
        for row, s in enumerate(self.summary, 2):
            ws.cell(row=row, column=1, value=s.get("type", "")).border = self.BORDER
            ws.cell(row=row, column=2, value=s.get("count", 0)).border = self.BORDER
            ws.cell(row=row, column=3, value=s.get("host_count", 0)).border = self.BORDER

        # 添加导出信息
        row = len(self.summary) + 4
        ws.cell(row=row, column=1, value="导出时间:")
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row + 1, column=1, value="凭证总数:")
        ws.cell(row=row + 1, column=2, value=len(self.credentials))